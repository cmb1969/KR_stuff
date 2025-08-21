import os
import re
import json
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import pdfplumber
import pandas as pd
from config import PDF_FIELDS, TABLE_CONFIG, PDF_SETTINGS

class PDFProcessor:
    def __init__(self, config: Dict[str, Any]):
        """Initialize the PDF processor with configuration."""
        self.config = config
        self.input_dir = Path(config["input_dir"])
        self.output_dir = Path(config["output_dir"])
        self.output_dir.mkdir(exist_ok=True)

    def extract_text_from_pdf(self, pdf_path: Path) -> str:
        """Extract all text from a PDF file."""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            print(f"Error extracting text from {pdf_path.name}: {str(e)}")
            return ""

    def extract_field_value(self, text: str, field_config: Dict[str, Any]) -> Any:
        """Extract a single field value from the text using regex or exact/variant labels.
        Supports:
        - field_config["label"]: single label
        - field_config["labels"]: list of alternative labels
        - field_config["group"]: regex capture group index when using regex (default 0)
        """
        label = field_config.get("label", "")
        labels = field_config.get("labels", ([label] if label else []))
        field_type = field_config.get("type", str)
        is_regex = field_config.get("is_regex", False)
        regex_group = int(field_config.get("group", 0))
        value_regex = field_config.get("value_regex")  # optional regex to locate value near label
        
        try:
            if is_regex:
                pattern = label
                match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                if match:
                    try:
                        value = match.group(regex_group).strip()
                    except IndexError:
                        value = match.group(0).strip()
                    print(f"Found (regex) {pattern}: {value}")
                    return value
                print(f"Warning: Could not match regex for {pattern}")
                return ""
            else:
                # Try each provided label variant
                label_variants = [lv for lv in labels if lv]
                if not label_variants and label:
                    label_variants = [label]
                
                for lbl in label_variants:
                    patterns = [
                        re.escape(lbl) + r"\s*[:\-]?\s*(.+?)(?:\n|$)",  # Label: Value
                        re.escape(lbl) + r"\s+(\S+)",                      # Label Value
                        re.escape(lbl) + r"\s*\n\s*(\S+)"                 # Label\nValue
                    ]
                    for pattern in patterns:
                        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                        if match:
                            value = match.group(1).strip()
                            # If a value_regex is provided, ensure the value matches; otherwise keep searching
                            if value_regex and not re.search(value_regex, value, re.IGNORECASE):
                                # Not acceptable, continue trying other patterns or fall back to window scan
                                continue
                            print(f"Found {lbl}: {value}")
                            try:
                                if field_type == int:
                                    value = re.sub(r"[^0-9.-]", "", value)
                                    return int(float(value)) if value else None
                                if field_type == float:
                                    value = re.sub(r"[^0-9.-]", "", value)
                                    return float(value) if value else None
                                return value
                            except (ValueError, TypeError):
                                return value
                    # If direct patterns failed, try scanning nearby lines after the label
                    # Find the first occurrence of the label and look ahead a few lines
                    lines = text.splitlines()
                    lbl_lower = lbl.lower()
                    for idx, line in enumerate(lines):
                        if lbl_lower in line.lower():
                            window_text = "\n".join(lines[idx + 1: idx + 8])  # look ahead up to 7 lines
                            # If a value_regex is provided, use it to find the value within the window
                            if value_regex:
                                m = re.search(value_regex, window_text, re.IGNORECASE)
                                if m:
                                    val = m.group(0).strip()
                                    print(f"Found near '{lbl}' using value_regex: {val}")
                                    try:
                                        if field_type == int:
                                            val = re.sub(r"[^0-9.-]", "", val)
                                            return int(float(val)) if val else None
                                        if field_type == float:
                                            val = re.sub(r"[^0-9.-]", "", val)
                                            return float(val) if val else None
                                        return val
                                    except (ValueError, TypeError):
                                        return val
                            # Otherwise, pick the first non-empty candidate line (apply value_regex if provided)
                            for cand in lines[idx + 1: idx + 8]:
                                cand = cand.strip()
                                if not cand:
                                    continue
                                if value_regex and not re.search(value_regex, cand, re.IGNORECASE):
                                    continue
                                print(f"Heuristic pick near '{lbl}': {cand}")
                                try:
                                    if field_type == int:
                                        cand_n = re.sub(r"[^0-9.-]", "", cand)
                                        return int(float(cand_n)) if cand_n else None
                                    if field_type == float:
                                        cand_n = re.sub(r"[^0-9.-]", "", cand)
                                        return float(cand_n) if cand_n else None
                                    return cand
                                except (ValueError, TypeError):
                                    return cand
                            break
                print(f"Warning: Could not find value for any of labels: {label_variants}")
                return ""
        except Exception as e:
            print(f"Error extracting field with labels {labels or label}: {str(e)}")
            return ""

    def extract_table_data(self, text: str) -> List[Dict[str, Any]]:
        """Extract table data from the text using case-insensitive markers and configured headers if provided."""
        table_data: List[Dict[str, Any]] = []
        lines = [line for line in text.split('\n')]  # Keep empty lines for better parsing
        lower_lines = [line.lower() for line in lines]
        start_marker = (TABLE_CONFIG.get("table_start") or "").lower()
        end_marker = (TABLE_CONFIG.get("table_end") or "").lower()
        
        # Find the start of the table (case-insensitive)
        start_index = -1
        for i, (line, ll) in enumerate(zip(lines, lower_lines)):
            if start_marker and start_marker in ll:
                start_index = i
                print(f"Detected table start at line {i}: {line}")
                break
        
        if start_index == -1:
            print(f"Warning: Could not find table start marker '{TABLE_CONFIG.get('table_start')}'")
            return table_data
        
        # Determine headers
        headers_cfg = TABLE_CONFIG.get("table_headers") or []
        headers: List[str] = []
        if headers_cfg:
            headers = headers_cfg
            # Assume the header row is right at or after start_index; advance one line for data start
            data_start = start_index + 1
        else:
            # Try to infer headers from the next few non-empty lines (allow extra spacing above headers)
            data_start = start_index + 1
            for i in range(start_index, min(start_index + 15, len(lines))):
                row = lines[i].strip()
                if not row:
                    continue
                potential_headers = [h.strip() for h in re.split(r'\s{2,}', row) if h.strip()]
                if len(potential_headers) >= 3:
                    headers = potential_headers
                    data_start = i + 1
                    break
        
        if not headers:
            print("Warning: Could not determine table headers")
            return table_data
        
        print(f"Using table headers: {headers}")
        
        # Extract data rows until end marker (case-insensitive)
        for raw_line, ll in zip(lines[data_start:], lower_lines[data_start:]):
            if end_marker and end_marker in ll:
                print(f"Detected table end at line: {raw_line}")
                break
            line = raw_line.strip()
            if not line or len(line.split()) < 2:
                continue
            values = [v.strip() for v in re.split(r'\s{2,}', line) if v.strip()]
            if not values:
                continue
            # Pad/truncate to match headers length
            if len(values) < len(headers):
                values = values + [''] * (len(headers) - len(values))
            elif len(values) > len(headers):
                values = values[:len(headers)]
            row_dict = dict(zip(headers, values))
            table_data.append(row_dict)
        
        print(f"Extracted {len(table_data)} rows from table")
        return table_data

    def extract_table_data_plumber(self, pdf_path: Path, anchor_after_text: Optional[str] = None) -> List[Dict[str, Any]]:
        """Extract table data using pdfplumber's table detection.
        Strategy:
        - Normalize header cells (collapse whitespace/newlines).
        - Score each table by overlap with TABLE_CONFIG['expected_headers'].
        - Pick best scoring table above TABLE_CONFIG['min_header_matches'].
        - If anchor_after_text is provided, only consider tables on or after the page containing that text.
        """
        results: List[Dict[str, Any]] = []
        expected = [h.strip().lower() for h in TABLE_CONFIG.get('expected_headers', [])]
        min_matches = int(TABLE_CONFIG.get('min_header_matches', max(1, len(expected)//3 or 1)))

        def norm_cell(s: Any) -> str:
            txt = (s or "")
            try:
                txt = str(txt)
            except Exception:
                txt = ""
            # collapse whitespace/newlines
            txt = re.sub(r"\s+", " ", txt).strip()
            # fix known header splits
            txt = txt.replace("PO O creation date", "PO creation date")
            return txt

        def header_score(headers: List[str]) -> int:
            h_low = [h.lower() for h in headers]
            score = 0
            for exp in expected:
                if any(exp in h for h in h_low):
                    score += 1
            return score

        best = {"score": -1, "headers": None, "rows": None, "page": None}
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Determine start page based on anchor text (e.g., coupon description value)
                start_page_idx = 0
                if anchor_after_text:
                    for p_idx, p in enumerate(pdf.pages):
                        try:
                            p_text = p.extract_text() or ""
                        except Exception:
                            p_text = ""
                        if anchor_after_text in p_text:
                            start_page_idx = p_idx
                            break
                if anchor_after_text:
                    print(f"pdfplumber: limiting table search to pages >= {start_page_idx + 1} due to anchor text match")
                for page_idx, page in enumerate(pdf.pages):
                    if page_idx < start_page_idx:
                        continue
                    tables = page.extract_tables() or []
                    for t_idx, table in enumerate(tables):
                        if not table or not any(table):
                            continue
                        # Determine header row: first non-empty row
                        header_row = None
                        start_row_idx = 0
                        for r_idx, row in enumerate(table):
                            if row and any(cell and str(cell).strip() for cell in row):
                                header_row = [ norm_cell(cell) for cell in row ]
                                start_row_idx = r_idx + 1
                                break
                        if not header_row:
                            continue
                        score = header_score(header_row)
                        # Build rows for this candidate
                        candidate_rows: List[Dict[str, Any]] = []
                        headers = header_row
                        for row in table[start_row_idx:]:
                            if not row:
                                continue
                            values = [ norm_cell(cell) for cell in row ]
                            if not any(values):
                                continue
                            # Pad/truncate to headers
                            if len(values) < len(headers):
                                values += [''] * (len(headers) - len(values))
                            elif len(values) > len(headers):
                                values = values[:len(headers)]
                            candidate_rows.append(dict(zip(headers, values)))
                        # Heuristics: require reasonable columns and at least 2 data rows
                        if len(headers) < 5 or len(candidate_rows) < 2:
                            continue
                        if score > best["score"]:
                            best = {
                                "score": score,
                                "headers": headers,
                                "rows": candidate_rows,
                                "page": page_idx + 1,
                            }
                if best["headers"] and best["score"] >= min_matches and best["rows"]:
                    print(f"pdfplumber: picked table on page {best['page']} with score {best['score']} / {len(expected)}; cols={len(best['headers'])}, rows={len(best['rows'])}; headers: {best['headers']}")
                    return best["rows"]
        except Exception as e:
            print(f"pdfplumber table extraction error: {e}")
        return results

    def process_pdf(self, pdf_path: Path) -> Dict[str, Any]:
        """Process a single PDF file and return extracted data."""
        print(f"Processing {pdf_path.name}...")
        text = self.extract_text_from_pdf(pdf_path)
        
        # Extract fields
        extracted_data = {}
        for field_name, field_config in PDF_FIELDS.items():
            value = self.extract_field_value(text, field_config)
            extracted_data[field_name] = value
        
        # Extract table data if needed
        if TABLE_CONFIG:
            # Prefer pdfplumber table extraction when possible
            section_anchor = TABLE_CONFIG.get("section_anchor")
            if section_anchor and isinstance(section_anchor, str) and section_anchor.strip():
                anchor_text = section_anchor.strip()
            else:
                anchor_text = extracted_data.get("coupon_description") if isinstance(extracted_data.get("coupon_description"), str) else None
            table_data = self.extract_table_data_plumber(pdf_path, anchor_after_text=anchor_text)
            if not table_data:
                table_data = self.extract_table_data(text)
            extracted_data["items"] = table_data
        
        return extracted_data

    def save_results(self, data: Dict[str, Any], filename: str):
        """Format the extracted data for Excel output."""
        # Create a list to hold all rows for this sheet
        sheet_data = []
        
        # Add invoice details as separate rows
        sheet_data.append(["Invoice Number", data.get('invoice_number', '')])
        sheet_data.append(["Coupon Description", data.get('coupon_description', '')])
        sheet_data.append(["Campaign Description", data.get('campaign_description', '')])
        sheet_data.append([])  # Empty row for spacing
        
        # Add table headers if there are items
        if 'items' in data and data['items']:
            # Get the headers from the first item
            headers = list(data['items'][0].keys())
            sheet_data.append(headers)
            
            # Add all data rows
            for item in data['items']:
                sheet_data.append([item.get(header, '') for header in headers])
            print(f"save_results: writing {len(data['items'])} table rows with headers: {headers}")
        
        return sheet_data

    def process_all_pdfs(self):
        """Process all PDF files in the input directory and save results."""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        
        pdf_files = list(self.input_dir.glob("*.pdf"))
        
        if not pdf_files:
            print(f"No PDF files found in {self.input_dir}")
            return
            
        print(f"Found {len(pdf_files)} PDF files to process.")
        
        # Create a new Excel workbook
        output_path = self.output_dir / self.config.get('output_filename', 'kroger_data.xlsx')
        wb = Workbook()
        
        # Remove the default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        for pdf_file in pdf_files:
            try:
                print(f"Processing {pdf_file.name}...")
                data = self.process_pdf(pdf_file)
                
                # Create a sheet for this PDF (use a shortened name if needed)
                sheet_name = pdf_file.stem[:31]  # Excel sheet names max 31 chars
                ws = wb.create_sheet(title=sheet_name)
                
                # Get the formatted data for this PDF
                sheet_data = self.save_results(data, pdf_file.stem)
                
                # Detect the header row index in sheet_data
                header_row_idx = None
                for idx, row in enumerate(sheet_data, 1):
                    if isinstance(row, list) and any(isinstance(c, str) and c.strip().lower() == 'line no' for c in row):
                        header_row_idx = idx
                        break
                if header_row_idx is None:
                    header_row_idx = 5  # default when we include 3 meta rows + blank

                # Write data to the worksheet
                for row_idx, row in enumerate(sheet_data, 1):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        # Style the header row
                        if row_idx == header_row_idx:
                            cell.font = Font(bold=True)

                # Log rows written to this sheet based on detected header row
                data_rows_written = max(0, len(sheet_data) - header_row_idx)
                print(f"Excel: wrote {data_rows_written} data rows to sheet '{sheet_name}'")
                
                # Auto-adjust column widths
                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
                
            except Exception as e:
                print(f"Error processing {pdf_file.name}: {str(e)}")
        
        # Save the workbook
        if len(wb.sheetnames) > 0:
            wb.save(output_path)
            print(f"\nAll data has been saved to: {output_path}")
        else:
            print("No data was extracted from any PDFs.")


def main():
    """Main function to run the PDF processor."""
    try:
        processor = PDFProcessor(PDF_SETTINGS)
        processor.process_all_pdfs()
        print("PDF processing completed successfully!")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        raise

if __name__ == "__main__":
    main()
